#!/usr/bin/env python3
"""Builds suggestions/WG7TF2-suggestions-sheet.xlsx from suggestions/sheet-plan.json.

The file is meant to be imported into the shared WG7-TF2 spreadsheet, so that
the no-account channel asks the same questions as the issue form, offers the
same closed lists, and states the same contribution terms.

Validation points at ranges on a generated lists sheet rather than at literal
lists, so refreshing the vocabulary is a paste into that sheet instead of
reopening every rule. Run: python3 scripts/build-sheet.py
"""

import json
import pathlib
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

ROOT = pathlib.Path(__file__).resolve().parent.parent
plan = json.loads((ROOT / "suggestions" / "sheet-plan.json").read_text())

ROWS = 300  # rows of the input sheet that carry validation
BRAND = "6D4AFF"

wb = Workbook()

# --- the lists, first, so validation can reference them --------------------
lists_ws = wb.active
lists_ws.title = plan["listsSheetName"]
lists_ws.sheet_state = "visible"  # hiding it makes the ranges harder to refresh
lists_ws["A1"] = (
    "Generated from the taxonomy by scripts/generate-sheet.mjs and "
    "scripts/build-sheet.py. Do not edit by hand: replace the columns from "
    "suggestions/options.csv in the repository, and the dropdowns follow."
)
lists_ws["A1"].font = Font(italic=True, size=9)

ranges = {}
for i, (name, values) in enumerate(plan["lists"].items(), start=1):
    col = get_column_letter(i)
    lists_ws.cell(row=2, column=i, value=name).font = Font(bold=True)
    for j, v in enumerate(values, start=3):
        lists_ws.cell(row=j, column=i, value=v)
    ranges[name] = f"'{plan['listsSheetName']}'!${col}$3:${col}${2 + len(values)}"
    lists_ws.column_dimensions[col].width = 46

# --- the input sheet -------------------------------------------------------
ws = wb.create_sheet(plan["sheetName"])
columns = plan["columns"]
last_col = get_column_letter(len(columns))

ws["A1"] = plan["notice"]
ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")
ws["A1"].font = Font(size=10)
ws.merge_cells(f"A1:{last_col}1")
ws.row_dimensions[1].height = 46

for i, c in enumerate(columns, start=1):
    cell = ws.cell(row=2, column=i, value=c["header"])
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=BRAND)
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.column_dimensions[get_column_letter(i)].width = 20 if c["list"] is None else 34
ws.row_dimensions[2].height = 30
ws.freeze_panes = "A3"

for i, c in enumerate(columns, start=1):
    if not c["list"]:
        continue
    col = get_column_letter(i)
    dv = DataValidation(
        type="list",
        formula1=ranges[c["list"]],
        allow_blank=True,
        showDropDown=False,  # openpyxl inverts this: False means show the arrow
        showErrorMessage=True,
        errorTitle="Not on the list",
        error=(
            "Pick one of the offered values. They come from the WG7-TF2 taxonomy, "
            "so answers can be matched with the ones filed through GitHub."
        ),
    )
    ws.add_data_validation(dv)
    dv.add(f"{col}3:{col}{ROWS}")

out = ROOT / "suggestions" / "WG7TF2-suggestions-sheet.xlsx"
wb.save(out)
print(
    f"Wrote {out.relative_to(ROOT)}: "
    f"{len(columns)} columns, "
    f"{sum(1 for c in columns if c['list'])} validated, "
    f"{len(plan['lists'])} lists"
)
